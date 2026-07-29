"""Norme de rapport CarburFlow — alignée sur les fiches de suivi terrain."""

from __future__ import annotations

import csv
import io
import unicodedata
from datetime import date, datetime, timedelta
from typing import Any

from django.db import transaction

from dashboard.models import Rapport

# Colonnes du modèle téléchargeable = fiche de suivi + période.
# Ordre et libellés volontairement proches de data/ligne_rapport.csv.
NORME_COLUMNS = [
    'date_debut',
    'date_fin',
    'id_cuve_principale',
    'id_cuve_journaliere',
    'id_groupe',
    'quantités_cuve_principale',
    'quantite_cuve_journaliere',
    'depotage',
    'compteur_horaire',
    'état_fonctionnement',
    'observations',
]

# Colonne utilisée pour l’export d’un rapport déjà importé, afin de refléter
# exactement la forme de la fiche de saisie / modèle initial sans les dates.
EXPORT_COLUMNS = [
    'id_cuve_journaliere',
    'site',
    'groupe_marque',
    'quantite_cuve_principale',
    'quantite_cuve_journaliere',
    'depotage',
    'compteur_horaire',
    'etat_fonctionnement',
    'observations',
]

# Alias acceptés à la lecture (anciens modèles / variantes sans accents).
COLUMN_ALIASES = {
    'code_cuve_journaliere': 'id_cuve_journaliere',
    'code cuve (verrouille)': 'id_cuve_journaliere',
    'code cuve (verrouillee)': 'id_cuve_journaliere',
    'code_cj': 'id_cuve_journaliere',
    'code_cuve': 'id_cuve_journaliere',
    'code': 'id_cuve_journaliere',
    'cuve journaliere (verrouille)': 'id_cuve_journaliere',
    'cuve journaliere (verrouillee)': 'id_cuve_journaliere',
    'cuve journaliere (laisser vide)': 'id_cuve_journaliere',
    'site / cuve principale': 'id_cuve_principale',
    'nom du nouveau site': 'id_cuve_principale',
    'groupe électrogène': 'id_groupe',
    'groupe electrogene': 'id_groupe',
    'quantité cp (l)': 'quantités_cuve_principale',
    'quantite cp (l)': 'quantités_cuve_principale',
    'quantité cj (l)': 'quantite_cuve_journaliere',
    'nom_site': 'id_cuve_principale',
    'nom_site_saisi': 'id_cuve_principale',
    'site': 'id_cuve_principale',
    'groupe / marque': 'id_groupe',
    'groupe_marque': 'id_groupe',
    'groupe': 'id_groupe',
    'marque groupe': 'marque_groupe',
    'puissance groupe': 'puissance_groupe',
    'code site existant (opt.)': 'code_site_existant',
    'capacite cp (l)': 'capacite_cp',
    'capacité cp (l)': 'capacite_cp',
    'capacite cj (l)': 'capacite_cj',
    'capacité cj (l)': 'capacite_cj',
    'quantite_gasoil_cuve_principale': 'quantités_cuve_principale',
    'quantites_cuve_principale': 'quantités_cuve_principale',
    'quantite_cuve_principale': 'quantités_cuve_principale',
    'quantite_gasoil_cuve_journaliere': 'quantite_cuve_journaliere',
    'quantite cj (l)': 'quantite_cuve_journaliere',
    'depotage (l)': 'depotage',
    'compteur horaire': 'compteur_horaire',
    'etat (f/p/hs)': 'état_fonctionnement',
    'etat_fonctionnement': 'état_fonctionnement',
    'etat': 'état_fonctionnement',
}


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize('NFKD', text)
    return ''.join(ch for ch in normalized if not unicodedata.combining(ch))


def canonicalize_header(name: str) -> str:
    raw = str(name or '').strip().lower()
    if not raw:
        return ''
    if raw in COLUMN_ALIASES:
        return COLUMN_ALIASES[raw]
    # Tentative sans accents
    ascii_key = _strip_accents(raw)
    if ascii_key in COLUMN_ALIASES:
        return COLUMN_ALIASES[ascii_key]
    # Match canonique sans accents (ex. etat_fonctionnement ↔ état_fonctionnement)
    for col in NORME_COLUMNS:
        if _strip_accents(col) == ascii_key or col == raw:
            return col
    return raw


def normalize_row_keys(row: dict) -> dict:
    """Renomme les clés d’une ligne vers les noms de colonnes de la fiche."""
    out = {}
    for key, value in (row or {}).items():
        if key is None:
            continue
        canon = canonicalize_header(str(key))
        if not canon:
            continue
        # Première occurrence gagne (évite d’écraser une valeur déjà canonique)
        if canon not in out or (out[canon] in (None, '') and value not in (None, '')):
            out[canon] = value

    # Zone « nouveaux sites » : CJ = site si vide ; marque/puissance restent séparés
    # (l’identifiant G{n}-MARQUE-PUISSANCE est composé par l’appli à la création).
    site = out.get('id_cuve_principale')
    cj = out.get('id_cuve_journaliere')
    if (not cj or str(cj).strip() == '') and site and str(site).strip():
        out['id_cuve_journaliere'] = str(site).strip()

    return out


def _is_section_separator(values) -> bool:
    first = str(values[0] if values and len(values) > 0 else '').strip()
    return first.startswith('---') or 'CARBURFLOW' in first.upper()


def _is_header_row(values) -> bool:
    """Détecte une ligne d’en-têtes (relevés ou zone nouveaux sites)."""
    cells = [str(v or '').strip().lower() for v in (values or []) if v is not None and str(v).strip()]
    if not cells:
        return False
    joined = ' | '.join(cells)
    markers = (
        'cuve journal',
        'nouveau site',
        'site / cuve',
        'groupe',
        'quantite',
        'quantité',
        'depotage',
        'dépotage',
        'observations',
    )
    hits = sum(1 for m in markers if m in joined)
    return hits >= 2


NORME_META = {
    'format': 'CarburFlow Fiche de suivi v1',
    'description': (
        'Même tableau que les fiches de suivi terrain. '
        'id_cuve_principale = nom du site (une cuve principale = un site). '
        'id_cuve_journaliere = nom de la cuve journalière. '
        'Ajoutez date_debut / date_fin (identique sur toutes les lignes), puis importez.'
    ),
    'columns': [
        {
            'name': 'date_debut',
            'label': 'Date de début',
            'type': 'date',
            'required': True,
            'example': '13/07/2026',
            'help': 'Début de la période du relevé. Identique sur toutes les lignes.',
        },
        {
            'name': 'date_fin',
            'label': 'Date de fin',
            'type': 'date',
            'required': True,
            'example': '17/07/2026',
            'help': 'Fin de la période du relevé. Identique sur toutes les lignes.',
        },
        {
            'name': 'id_cuve_principale',
            'label': 'Cuve principale (site)',
            'type': 'string',
            'required': False,
            'example': 'BEPANDA INTERNATIONAL',
            'help': 'Nom du site / cuve principale, comme sur la fiche de suivi.',
        },
        {
            'name': 'id_cuve_journaliere',
            'label': 'Cuve journalière',
            'type': 'string',
            'required': False,
            'example': 'BEPANDA INTERNATIONAL',
            'help': 'Nom de la cuve journalière (fiche de suivi).',
        },
        {
            'name': 'id_groupe',
            'label': 'N° groupe électrogène',
            'type': 'int',
            'required': False,
            'example': '1',
            'help': 'Numéro du groupe électrogène (comme sur la fiche).',
        },
        {
            'name': 'quantités_cuve_principale',
            'label': 'Quantité cuve principale (L)',
            'type': 'float',
            'required': False,
            'example': '8448',
            'help': 'Litres mesurés dans la cuve principale.',
        },
        {
            'name': 'quantite_cuve_journaliere',
            'label': 'Quantité cuve journalière (L)',
            'type': 'float',
            'required': False,
            'example': '1000',
            'help': 'Litres mesurés dans la cuve journalière.',
        },
        {
            'name': 'depotage',
            'label': 'Dépotage (L)',
            'type': 'float',
            'required': False,
            'example': '0',
            'help': 'Volume dépoté sur la période (0 si aucun).',
        },
        {
            'name': 'compteur_horaire',
            'label': 'Compteur horaire',
            'type': 'float',
            'required': False,
            'example': '1864',
            'help': 'Relevé du compteur horaire du groupe.',
        },
        {
            'name': 'état_fonctionnement',
            'label': 'État de fonctionnement',
            'type': 'string',
            'required': False,
            'example': 'F',
            'help': 'Code d’état (ex. F = fonctionnement).',
        },
        {
            'name': 'observations',
            'label': 'Observations',
            'type': 'string',
            'required': False,
            'example': 'RAS',
            'help': 'Commentaire libre (incident, maintenance, RAS…).',
        },
    ],
}

SAMPLE_ROWS = [
    {
        'date_debut': '13/07/2026',
        'date_fin': '17/07/2026',
        'id_cuve_principale': 'BEPANDA INTERNATIONAL',
        'id_cuve_journaliere': 'BEPANDA INTERNATIONAL',
        'id_groupe': '1',
        'quantités_cuve_principale': '8448',
        'quantite_cuve_journaliere': '1000',
        'depotage': '0',
        'compteur_horaire': '1864',
        'état_fonctionnement': 'F',
        'observations': 'RAS',
    },
]


def build_csv_bytes(include_sample: bool = True) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=NORME_COLUMNS, lineterminator='\n')
    writer.writeheader()
    if include_sample:
        for row in SAMPLE_ROWS:
            writer.writerow(row)
    return buffer.getvalue().encode('utf-8-sig')


def build_xlsx_bytes(include_sample: bool = True) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rapport'
    ws.append(NORME_COLUMNS)
    if include_sample:
        for row in SAMPLE_ROWS:
            ws.append([row.get(col, '') for col in NORME_COLUMNS])

    meta = wb.create_sheet('Meta')
    meta.append(['champ', 'valeur'])
    meta.append(['format', NORME_META['format']])
    meta.append(['description', NORME_META['description']])
    meta.append(['colonnes', ', '.join(NORME_COLUMNS)])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def rapport_to_rows(rapport: Rapport) -> list[dict]:
    """Reconstitue les lignes au format de la fiche de saisie pour export."""
    rows = []
    lignes = rapport.lignes.select_related(
        'cuve_principale',
        'cuve_journaliere',
        'groupe_electrogene',
    ).all()
    for ligne in lignes:
        groupe_display = ''
        if ligne.groupe_electrogene_id:
            ge = ligne.groupe_electrogene
            marque = getattr(ge, 'marque', None)
            groupe_display = (
                f'{ge.identifiant} ({marque})'
                if marque
                else ge.identifiant
            )

        rows.append({
            'id_cuve_journaliere': (
                ligne.cuve_journaliere.identifiant if ligne.cuve_journaliere_id else ''
            ),
            'site': (
                ligne.cuve_principale.identifiant if ligne.cuve_principale_id else ''
            ),
            'groupe_marque': groupe_display,
            'quantite_cuve_principale': ligne.quantite_gasoil_cuve_principale
            if ligne.quantite_gasoil_cuve_principale is not None
            else '',
            'quantite_cuve_journaliere': ligne.quantite_gasoil_cuve_journaliere
            if ligne.quantite_gasoil_cuve_journaliere is not None
            else '',
            'depotage': ligne.depotage if ligne.depotage is not None else '',
            'compteur_horaire': ligne.compteur_horaire if ligne.compteur_horaire is not None else '',
            'etat_fonctionnement': ligne.etat_fonctionnement or '',
            'observations': ligne.observations or '',
        })
    return rows


def build_rapport_csv_bytes(rapport: Rapport) -> bytes:
    buffer = io.StringIO()
    buffer.write('# CARBURFLOW — RAPPORT IMPORTÉ\n')
    buffer.write(f"# date_debut: {rapport.date_debut.strftime('%d/%m/%Y') if rapport.date_debut else ''}\n")
    buffer.write(f"# date_fin: {rapport.date_fin.strftime('%d/%m/%Y') if rapport.date_fin else ''}\n")
    buffer.write(f"# rapport_id: {rapport.id}\n")
    if rapport.created_by_id:
        buffer.write(f"# created_by: {rapport.created_by.get_username()}\n")
    buffer.write('# colonnes: ' + ', '.join(EXPORT_COLUMNS) + '\n')
    writer = csv.DictWriter(buffer, fieldnames=EXPORT_COLUMNS, lineterminator='\n')
    writer.writeheader()
    for row in rapport_to_rows(rapport):
        writer.writerow({col: row.get(col, '') for col in EXPORT_COLUMNS})
    return buffer.getvalue().encode('utf-8-sig')


def build_rapport_xlsx_bytes(rapport: Rapport) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = 'Entête'
    ws_meta.append(['CARBURFLOW — RAPPORT IMPORTÉ'])
    ws_meta.append([])
    ws_meta.append(['champ', 'valeur'])
    ws_meta.append(['rapport_id', rapport.id])
    ws_meta.append(['date_debut', str(rapport.date_debut)])
    ws_meta.append(['date_fin', str(rapport.date_fin)])
    if rapport.created_by_id:
        ws_meta.append(['created_by', rapport.created_by.get_username()])

    ws = wb.create_sheet(title='Relevés')
    ws.append(EXPORT_COLUMNS)
    for row in rapport_to_rows(rapport):
        ws.append([row.get(col, '') for col in EXPORT_COLUMNS])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


COLUMN_LABELS = {
    col['name']: col.get('label') or col['name'] for col in NORME_META['columns']
}


class ImportValidationError(Exception):
    """Erreur d'import avec détails compréhensibles (ligne / colonne)."""

    def __init__(self, message: str, errors: list[dict] | None = None):
        super().__init__(message)
        self.message = message
        self.errors = errors or []

    def as_dict(self) -> dict:
        return {
            'detail': self.message,
            'errors': self.errors,
        }


def _friendly_error(
    *,
    row: int | None,
    column: str | None,
    message: str,
    how_to_fix: str,
) -> dict:
    return {
        'row': row,
        'column': column,
        'column_label': COLUMN_LABELS.get(column or '', column),
        'message': message,
        'how_to_fix': how_to_fix,
    }


def _parse_date(value: Any, *, row: int | None = None, column: str = 'date') -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Numéro de série Excel (ex. 46237 = 03/08/2026) — évite les ambiguïtés locale jj/mm
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if 20000 <= serial <= 80000:  # ~1954 → ~2119
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    label = COLUMN_LABELS.get(column, column)
    raise ImportValidationError(
        f'Une date n’est pas au bon format (ligne {row}).',
        [
            _friendly_error(
                row=row,
                column=column,
                message=f'La valeur « {text} » n’est pas une date reconnue pour « {label} ».',
                how_to_fix='Écrivez la date comme ceci : 13/07/2026 ou 2026-07-13.',
            )
        ],
    )


def _swap_day_month(value: date) -> date | None:
    """Échange jj↔mm si les deux sont ≤ 12 (ambiguïté Excel US vs FR)."""
    if value.day > 12 or value.month > 12:
        return None
    if value.day == value.month:
        return None
    try:
        return date(value.year, value.day, value.month)
    except ValueError:
        return None


def coerce_french_week_period(
    d1: date | None,
    d2: date | None,
    *,
    max_days: int = 14,
) -> tuple[date | None, date | None, bool]:
    """
    Corrige le cas Excel US : 03/08/2026 (3 août) lu comme 08/03/2026 (8 mars).

    Si la période brute dépasse max_days, tente d’échanger jj/mm sur date_debut
    (puis sur les deux dates) pour retrouver une semaine plausible.
    """
    if d1 is None or d2 is None:
        return d1, d2, False
    if d1 <= d2 and (d2 - d1).days <= max_days:
        return d1, d2, False

    candidates: list[tuple[int, date, date]] = []
    swapped_debut = _swap_day_month(d1)
    swapped_fin = _swap_day_month(d2)

    if swapped_debut is not None:
        if swapped_debut <= d2 and (d2 - swapped_debut).days <= max_days:
            candidates.append((1, swapped_debut, d2))
        if swapped_fin is not None and swapped_debut <= swapped_fin and (swapped_fin - swapped_debut).days <= max_days:
            candidates.append((2, swapped_debut, swapped_fin))

    if not candidates:
        return d1, d2, False

    candidates.sort(key=lambda item: (item[0], (item[2] - item[1]).days))
    _, fixed_debut, fixed_fin = candidates[0]
    return fixed_debut, fixed_fin, True


def _to_float(value: Any, default: float = 0.0, *, row: int | None = None, column: str | None = None) -> float:
    if value is None or value == '':
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '.').replace(' ', '')
    if not text:
        return default
    try:
        return float(text)
    except ValueError as exc:
        label = COLUMN_LABELS.get(column or '', column or 'nombre')
        raise ImportValidationError(
            f'Un nombre est incorrect (ligne {row}).',
            [
                _friendly_error(
                    row=row,
                    column=column,
                    message=f'La valeur « {value} » n’est pas un nombre valide pour « {label} ».',
                    how_to_fix='Mettez uniquement un nombre, par exemple 4500 ou 12,5 (sans lettres).',
                )
            ],
        ) from exc


def _to_int_or_none(value: Any, *, row: int | None = None, column: str | None = None) -> int | None:
    if value is None or value == '':
        return None
    try:
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if not text:
            return None
        return int(float(text))
    except ValueError as exc:
        label = COLUMN_LABELS.get(column or '', column or 'identifiant')
        raise ImportValidationError(
            f'Un numéro est incorrect (ligne {row}).',
            [
                _friendly_error(
                    row=row,
                    column=column,
                    message=f'La valeur « {value} » n’est pas un numéro valide pour « {label} ».',
                    how_to_fix='Mettez uniquement le numéro du groupe (chiffres, sans lettres).',
                )
            ],
        ) from exc


def _to_name_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_headers(headers: list[str]) -> list[str]:
    return [canonicalize_header(h) for h in headers]


def _missing_required_columns(headers: list[str]) -> list[dict]:
    missing = [c for c in ('date_debut', 'date_fin') if c not in headers]
    if not missing:
        return []
    return [
        _friendly_error(
            row=1,
            column=col,
            message=f'La colonne « {COLUMN_LABELS.get(col, col)} » est absente du fichier.',
            how_to_fix='Retéléchargez le modèle (étape 1) et ne renommez pas les titres des colonnes.',
        )
        for col in missing
    ]


def _detect_csv_delimiter(sample: str) -> str:
    """Détecte ; ou , (Excel FR utilise souvent ;)."""
    lines = [line.strip() for line in sample.splitlines() if line.strip()]
    candidate_lines = [line for line in lines if not line.startswith('#')]
    if not candidate_lines:
        return ','
    first_line = candidate_lines[0]
    if first_line.count(';') > first_line.count(','):
        return ';'
    return ','


def _extract_csv_meta(text: str) -> dict:
    meta: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line.startswith('#'):
            continue
        if ':' not in line:
            continue
        key, value = line[1:].split(':', 1)
        key = canonicalize_header(key.strip())
        if key in {'date_debut', 'date_fin', 'rapport_id', 'created_by', 'colonnes'}:
            meta[key] = value.strip()
    return meta


def rows_from_csv(file_bytes: bytes) -> list[dict]:
    try:
        text = file_bytes.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ImportValidationError(
            'Le fichier CSV n’est pas lisible.',
            [
                _friendly_error(
                    row=None,
                    column=None,
                    message='Le fichier semble mal enregistré (encodage incorrect).',
                    how_to_fix='Enregistrez le fichier en CSV UTF-8, ou utilisez plutôt le modèle Excel.',
                )
            ],
        ) from exc
    meta_info = _extract_csv_meta(text)
    delimiter = _detect_csv_delimiter(text)
    lines = [line for line in text.splitlines() if not line.lstrip().startswith('#')]
    cleaned_text = '\n'.join(lines)
    reader = csv.DictReader(io.StringIO(cleaned_text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ImportValidationError(
            'Le fichier CSV est vide ou sans titres de colonnes.',
            [
                _friendly_error(
                    row=1,
                    column=None,
                    message='La première ligne (titres) est manquante.',
                    how_to_fix='Utilisez le modèle téléchargé à l’étape 1 : la 1re ligne doit contenir les noms des colonnes.',
                )
            ],
        )
    headers = _normalize_headers(reader.fieldnames)
    if 'date_debut' in meta_info and 'date_debut' not in headers:
        headers.append('date_debut')
    if 'date_fin' in meta_info and 'date_fin' not in headers:
        headers.append('date_fin')
    missing_errors = _missing_required_columns(headers)
    if missing_errors:
        raise ImportValidationError(
            'Il manque des colonnes obligatoires dans votre fichier.',
            missing_errors,
        )
    rows = []
    for raw in reader:
        row = normalize_row_keys({k: v for k, v in raw.items() if k})
        # Meta CSV (# date_debut: ...) = seule source de vérité pour la période
        if 'date_debut' in meta_info:
            row['date_debut'] = meta_info['date_debut']
        if 'date_fin' in meta_info:
            row['date_fin'] = meta_info['date_fin']
        if not any(str(v or '').strip() for v in row.values()):
            continue
        rows.append(row)
    return rows


def rows_from_xlsx(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ImportValidationError(
            'Le fichier Excel n’a pas pu être ouvert.',
            [
                _friendly_error(
                    row=None,
                    column=None,
                    message='Le fichier est peut-être endommagé ou n’est pas un vrai Excel.',
                    how_to_fix='Retéléchargez le modèle Excel, recopiez vos données dedans, puis réessayez.',
                )
            ],
        ) from exc

    meta_info = {}
    # Extrait les métadonnées de la feuille Entête / Meta si présente
    for meta_sheet_name in ('Entête', 'Entete', 'Meta'):
        if meta_sheet_name in wb.sheetnames:
            ws_meta = wb[meta_sheet_name]
            for row in ws_meta.iter_rows(values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    k = canonicalize_header(str(row[0]))
                    if k:
                        meta_info[k] = row[1]
            break

    # Sélectionne la feuille de données (Relevés, Rapport ou la première feuille si non trouvée)
    data_sheet_name = None
    for target in ('Relevés', 'Releves', 'Rapport'):
        if target in wb.sheetnames:
            data_sheet_name = target
            break
    if not data_sheet_name:
        non_meta = [name for name in wb.sheetnames if name not in ('Entête', 'Entete', 'Meta')]
        data_sheet_name = non_meta[0] if non_meta else wb.sheetnames[0]

    ws = wb[data_sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    headers = []
    start_idx = 0
    for i, row_vals in enumerate(all_rows):
        if not row_vals or all(v is None or str(v).strip() == '' for v in row_vals):
            continue
        if _is_section_separator(row_vals):
            continue
        headers = _normalize_headers(list(row_vals))
        start_idx = i + 1
        break

    if not headers:
        raise ImportValidationError(
            'Le fichier Excel est vide.',
            [
                _friendly_error(
                    row=1,
                    column=None,
                    message='Aucune ligne d’en-tête trouvée dans la feuille de relevés.',
                    how_to_fix='Utilisez la fiche de relevé hebdomadaire générée.',
                )
            ],
        )

    if 'date_debut' in meta_info and 'date_debut' not in headers:
        headers.append('date_debut')
    if 'date_fin' in meta_info and 'date_fin' not in headers:
        headers.append('date_fin')

    missing_errors = _missing_required_columns(headers)
    if missing_errors:
        raise ImportValidationError(
            'Il manque des colonnes obligatoires dans votre fichier.',
            missing_errors,
        )

    rows = []
    for values in all_rows[start_idx:]:
        if values is None or all(v is None or str(v).strip() == '' for v in values):
            continue
        if _is_section_separator(values):
            continue
        # Nouvelle zone = nouveaux en-têtes (ex. « Nom du Nouveau Site ») :
        # basculer le mapping au lieu de réutiliser les colonnes des anciens relevés.
        if _is_header_row(values):
            headers = _normalize_headers(list(values))
            if 'date_debut' in meta_info and 'date_debut' not in headers:
                headers.append('date_debut')
            if 'date_fin' in meta_info and 'date_fin' not in headers:
                headers.append('date_fin')
            continue

        raw = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            if idx < len(values):
                raw[key] = values[idx]
            elif key in meta_info:
                raw[key] = meta_info[key]
            else:
                raw[key] = None

        # Entête = seule source de vérité pour la période (jamais une colonne par ligne)
        if 'date_debut' in meta_info:
            raw['date_debut'] = meta_info['date_debut']
        if 'date_fin' in meta_info:
            raw['date_fin'] = meta_info['date_fin']

        normalized = normalize_row_keys(raw)
        if any(
            v is not None and str(v).strip() != ''
            for k, v in normalized.items()
            if k not in ('date_debut', 'date_fin', 'capacite_cp', 'capacite_cj', 'code_site_existant')
        ):
            rows.append(normalized)

    return rows


@transaction.atomic
def import_report_rows(rows: list[dict], user, *, create_missing: bool = False) -> tuple[Rapport, int]:
    """
    Import web / API : délègue au pipeline (analyse + insertion).
    Par défaut refuse les IDs inconnus (create_missing=False).
    """
    from dashboard.rapport_pipeline import import_rapport_lignes

    normalized = [normalize_row_keys(r) for r in rows]
    result = import_rapport_lignes(
        normalized,
        user=user,
        create_missing=create_missing,
        require_entities=True,
    )
    rapport = Rapport.objects.get(pk=result.rapport_id)
    return rapport, result.imported_lines
