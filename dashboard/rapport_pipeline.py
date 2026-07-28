"""
Pipeline d'import de rapports CarburFlow (3 étapes) — fiches de suivi.

1. analyze_rapport_* : cohérence + nouveaux sites / cuves journalières / groupes
2. create_entities_from_analysis : création des entités manquantes
3. import_rapport_lignes : Rapport + LigneRapport

Identifiants terrain :
- id_cuve_principale = nom du site (string), une cuve principale = un site
- id_cuve_journaliere = nom de la cuve journalière (string)
- id_groupe = numéro (int)
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import date as date_cls
from pathlib import Path

from django.db import transaction

from dashboard.models import (
    CuveJournaliere,
    CuvePrincipale,
    GroupeElectrogene,
    LigneRapport,
    Rapport,
)
from dashboard.norme import (
    ImportValidationError,
    _friendly_error,
    _parse_date,
    coerce_french_week_period,
    _to_float,
    _to_int_or_none,
    _to_name_or_none,
    normalize_row_keys,
    rows_from_csv,
    rows_from_xlsx,
)


@dataclass
class EntityRef:
    kind: str  # cuve_principale | cuve_journaliere | groupe
    key: str  # nom site / nom CJ / id groupe en str
    label: str
    exists: bool
    sample_rows: list[int] = field(default_factory=list)


@dataclass
class AnalysisIssue:
    level: str  # error | warning
    row: int | None
    column: str | None
    message: str
    how_to_fix: str = ''


@dataclass
class AnalysisResult:
    ok: bool
    row_count: int
    date_debut: str | None
    date_fin: str | None
    issues: list[AnalysisIssue] = field(default_factory=list)
    known_cuves_principales: list[EntityRef] = field(default_factory=list)
    known_cuves_journalieres: list[EntityRef] = field(default_factory=list)
    known_groupes: list[EntityRef] = field(default_factory=list)
    new_cuves_principales: list[EntityRef] = field(default_factory=list)
    new_cuves_journalieres: list[EntityRef] = field(default_factory=list)
    new_groupes: list[EntityRef] = field(default_factory=list)
    # Liens suggérés pour création CJ : nom_cj → {site, groupe_id}
    cj_links: dict[str, dict] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return asdict(self)

    @property
    def errors(self) -> list[AnalysisIssue]:
        return [i for i in self.issues if i.level == 'error']

    @property
    def warnings(self) -> list[AnalysisIssue]:
        return [i for i in self.issues if i.level == 'warning']


@dataclass
class CreateResult:
    created_cuves_principales: list[str] = field(default_factory=list)
    created_cuves_journalieres: list[str] = field(default_factory=list)
    created_groupes: list[str] = field(default_factory=list)
    skipped_existing: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ImportResult:
    rapport_id: int
    imported_lines: int
    date_debut: str
    date_fin: str
    created_entities: CreateResult | None = None

    def to_dict(self) -> dict:
        return asdict(self)


def load_rapport_rows(path: str | Path) -> list[dict]:
    path = Path(path)
    if not path.exists():
        raise ImportValidationError(
            f'Fichier introuvable : {path}',
            [
                _friendly_error(
                    row=None,
                    column=None,
                    message=f'Le fichier « {path} » n’existe pas.',
                    how_to_fix='Vérifiez le chemin du fichier .xlsx ou .csv.',
                )
            ],
        )
    raw = path.read_bytes()
    lower = path.name.lower()
    if lower.endswith('.xlsx'):
        return rows_from_xlsx(raw)
    if lower.endswith('.csv'):
        return rows_from_csv(raw)
    raise ImportValidationError(
        'Type de fichier non accepté.',
        [
            _friendly_error(
                row=None,
                column=None,
                message=f'Le fichier « {path.name} » n’est pas un Excel ou un CSV.',
                how_to_fix='Utilisez un fichier .xlsx ou .csv au format fiche de suivi.',
            )
        ],
    )


def load_rapport_rows_from_bytes(filename: str, raw: bytes) -> list[dict]:
    lower = (filename or '').lower()
    if lower.endswith('.xlsx'):
        return rows_from_xlsx(raw)
    if lower.endswith('.csv'):
        return rows_from_csv(raw)
    raise ImportValidationError(
        'Type de fichier non accepté.',
        [
            _friendly_error(
                row=None,
                column=None,
                message=f'Le fichier « {filename} » n’est pas un Excel ou un CSV.',
                how_to_fix='Choisissez un fichier .xlsx ou .csv (modèle étape 1).',
            )
        ],
    )


def _lookup_cp(name: str | None) -> CuvePrincipale | None:
    if not name:
        return None
    return CuvePrincipale.objects.filter(identifiant__iexact=name.strip()).first()


def _lookup_cj(name: str | None) -> CuveJournaliere | None:
    if not name:
        return None
    return CuveJournaliere.objects.filter(identifiant__iexact=name.strip()).first()


def _lookup_groupe(gid: int | None) -> GroupeElectrogene | None:
    if gid is None:
        return None
    return GroupeElectrogene.objects.filter(pk=gid).first()


def _lookup_groupe_by_identifiant(identifiant: str | None) -> GroupeElectrogene | None:
    if not identifiant:
        return None
    text = str(identifiant).strip()
    if not text:
        return None
    # Affichage fiche : "G1-SDMO-830 (SDMO)" → "G1-SDMO-830"
    if ' (' in text and text.endswith(')'):
        text = text[: text.rfind(' (')].strip()
    obj = GroupeElectrogene.objects.filter(identifiant__iexact=text).first()
    if obj:
        return obj
    gid = _extract_groupe_id(text)
    if gid is not None:
        return GroupeElectrogene.objects.filter(pk=gid).first()
    return None


def _groupe_key_from_raw(raw: object) -> str | None:
    """Identifiant métier du groupe (texte), sans le suffixe « (marque) »."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if ' (' in text and text.endswith(')'):
        text = text[: text.rfind(' (')].strip()
    return text or None


def _pending_groupe_key(marque: str | None, puissance: str | None) -> str:
    m = (marque or '').strip().upper()
    p = str(puissance or '').strip()
    return f'__NEW__{m}__{p}'


def _parse_pending_groupe_key(key: str) -> tuple[str, str] | None:
    if not key or not str(key).startswith('__NEW__'):
        return None
    rest = str(key)[len('__NEW__') :]
    if '__' not in rest:
        return (rest, '')
    marque, puissance = rest.split('__', 1)
    return marque, puissance


def _next_groupe_seq() -> int:
    """Prochain numéro n pour G{n}-MARQUE-PUISSANCE."""
    import re

    max_n = 0
    for ident in GroupeElectrogene.objects.values_list('identifiant', flat=True):
        match = re.match(r'^G(\d+)[-_]', str(ident or ''), re.IGNORECASE)
        if match:
            max_n = max(max_n, int(match.group(1)))
    return max_n + 1


def compose_groupe_identifiant(marque: str | None, puissance: str | None) -> str:
    """
    Construit l’identifiant condensé du groupe : G{n}-SDMO-830
    à partir de la marque et de la puissance (la numérotation est gérée par l’appli).
    """
    n = _next_groupe_seq()
    m = (marque or 'SANS').strip().upper() or 'SANS'
    p = str(puissance if puissance is not None and str(puissance).strip() != '' else '0').strip()
    return f'G{n}-{m}-{p}'


def _lookup_groupe_by_marque_puissance(
    marque: str | None,
    puissance: str | None,
) -> GroupeElectrogene | None:
    qs = GroupeElectrogene.objects.all()
    if marque and str(marque).strip():
        qs = qs.filter(marque__iexact=str(marque).strip())
    else:
        return None
    if puissance is not None and str(puissance).strip() != '':
        qs = qs.filter(puissance__iexact=str(puissance).strip())
    return qs.first()


def _extract_groupe_id(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    import re
    match = re.search(r'(\d+)', text)
    if match:
        return int(match.group(1))
    return None


def generate_rapport_template_xlsx(
    date_debut: date_cls | str | None = None,
    date_fin: date_cls | str | None = None,
) -> bytes:
    """
    Génère la fiche de relevé hebdomadaire au format Excel (.xlsx) pré-remplie
    avec les sites, cuves journalières et groupes existants en base.
    """
    import io
    from datetime import date as date_type, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    if isinstance(date_debut, str) and date_debut.strip():
        date_debut = _parse_date(date_debut)
    elif not isinstance(date_debut, date_cls):
        date_debut = None

    if isinstance(date_fin, str) and date_fin.strip():
        date_fin = _parse_date(date_fin)
    elif not isinstance(date_fin, date_cls):
        date_fin = None

    if not date_debut or not date_fin:
        today = date_type.today()
        date_debut = today - timedelta(days=today.weekday())
        date_fin = date_debut + timedelta(days=6)

    wb = Workbook()

    # Feuille 1 : Entête
    ws_meta = wb.active
    ws_meta.title = 'Entête'
    ws_meta.views.sheetView[0].showGridLines = True

    meta_title_font = Font(name='Arial', size=14, bold=True, color='0B3D7A')
    label_font = Font(name='Arial', size=10, bold=True)

    ws_meta.append(['CARBURFLOW — FICHE DE RELEVÉ HEBDOMADAIRE'])
    ws_meta.cell(row=1, column=1).font = meta_title_font
    ws_meta.append([])

    meta_rows = [
        ('date_debut', date_debut, 'Date de début du relevé (jj/mm/aaaa)'),
        ('date_fin', date_fin, 'Date de fin du relevé (jj/mm/aaaa)'),
        ('direction_regionale', 'DOUALA', 'Direction régionale'),
        ('centre', 'CENTRE LITTORAL', 'Centre / Zone'),
        ('responsable', '', 'Nom du responsable / opérateur'),
        ('date_generation', date_type.today(), 'Date de génération du fichier'),
    ]

    ws_meta.append(['Champ', 'Valeur', 'Description'])
    ws_meta.cell(row=3, column=1).font = label_font
    ws_meta.cell(row=3, column=2).font = label_font
    ws_meta.cell(row=3, column=3).font = label_font

    for item in meta_rows:
        ws_meta.append(list(item))

    # Dates en TEXTE jj/mm/aaaa : Excel ne peut plus les réinterpréter en mm/jj (locale US)
    for row_idx in (4, 5):  # date_debut, date_fin
        cell = ws_meta.cell(row=row_idx, column=2)
        if isinstance(cell.value, date_type):
            cell.value = cell.value.strftime('%d/%m/%Y')
        cell.number_format = '@'
    gen_cell = ws_meta.cell(row=9, column=2)
    if isinstance(gen_cell.value, date_type):
        gen_cell.number_format = 'DD/MM/YYYY'

    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 25
    ws_meta.column_dimensions['C'].width = 45

    # Feuille 2 : Relevés
    ws = wb.create_sheet(title='Relevés')
    ws.views.sheetView[0].showGridLines = True

    headers = [
        'id_cuve_journaliere',
        'site',
        'groupe_marque',
        'quantite_cuve_principale',
        'quantite_cuve_journaliere',
        'depotage',
        'compteur_horaire',
        'etat_fonctionnement',
        'observations',
    ]

    header_labels = [
        'Cuve journalière (Verrouillé)',
        'Site / Cuve Principale',
        'Groupe Électrogène',
        'Quantité CP (L)',
        'Quantité CJ (L)',
        'Dépotage (L)',
        'Compteur Horaire',
        'État (F/P/HS)',
        'Observations',
    ]

    ws.append(header_labels)

    header_fill = PatternFill(start_color='0B3D7A', end_color='0B3D7A', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    locked_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cuves = (
        CuveJournaliere.objects.select_related('cuve_principale', 'groupe_electrogene')
        .all()
        .order_by('cuve_principale__identifiant', 'identifiant', 'id')
    )

    current_row = 2
    for cj in cuves:
        cp_name = cj.cuve_principale.identifiant if cj.cuve_principale else '—'
        g_str = (
            f'{cj.groupe_electrogene.identifiant} ({cj.groupe_electrogene.marque})'
            if cj.groupe_electrogene
            else 'Sans groupe'
        )

        row_vals = [
            cj.identifiant,
            cp_name,
            g_str,
            '',
            '',
            0,
            '',
            'F',
            '',
        ]
        ws.append(row_vals)

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.border = thin_border
            if col_idx in (1, 2, 3):
                c.protection = Protection(locked=True)
                c.fill = locked_fill
            else:
                c.protection = Protection(locked=False)

        current_row += 1

    current_row += 1
    sep_cell = ws.cell(
        row=current_row,
        column=1,
        value=(
            '--- NOUVEAUX SITES / GROUPES — renseigner le site, la marque et la puissance '
            '(l’identifiant G{n}-MARQUE-PUISSANCE est généré automatiquement) ---'
        ),
    )
    sep_cell.font = Font(name='Arial', size=10, bold=True, color='1E3A8A')
    current_row += 1

    new_labels = [
        'Cuve journalière (optionnel)',
        'Nom du Nouveau Site',
        'Marque Groupe',
        'Puissance Groupe',
        'Capacité CP (L)',
        'Capacité CJ (L)',
        'Quantité CP (L)',
        'Quantité CJ (L)',
        'Dépotage (L)',
        'Compteur Horaire',
        'État (F/P/HS)',
        'Observations',
    ]
    new_header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    for col_idx, label in enumerate(new_labels, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=label)
        c.fill = new_header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1

    unlocked_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    for _ in range(5):
        for col_idx in range(1, len(new_labels) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.border = thin_border
            c.protection = Protection(locked=False)
            c.fill = unlocked_fill
        current_row += 1

    dv = DataValidation(type='list', formula1='"F,P,HS"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'H2:H{current_row}')
    # Colonne État de la zone nouveaux sites (K)
    dv.add(f'K{current_row - 5}:K{current_row}')

    ws.protection.sheet = True
    ws.protection.enable()

    column_widths = {
        'A': 28, 'B': 26, 'C': 16, 'D': 16, 'E': 14,
        'F': 14, 'G': 14, 'H': 14, 'I': 12, 'J': 16, 'K': 14, 'L': 30,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def analyze_rapport_rows(rows: list[dict], *, create_missing: bool = False) -> AnalysisResult:
    issues: list[AnalysisIssue] = []
    rows = [normalize_row_keys(r) for r in rows]

    if not rows:
        issues.append(
            AnalysisIssue(
                level='error',
                row=2,
                column=None,
                message='Aucune ligne de relevé dans le fichier.',
                how_to_fix='Ajoutez au moins une ligne sous les titres.',
            )
        )
        return AnalysisResult(ok=False, row_count=0, date_debut=None, date_fin=None, issues=issues)

    date_debut = None
    date_fin = None
    seen_keys: dict[tuple, int] = {}

    cp_names: dict[str, list[int]] = {}
    cj_names: dict[str, list[int]] = {}
    g_names: dict[str, list[int]] = {}
    cj_links: dict[str, dict] = {}

    # Période lue une seule fois (Entête / 1re valeur), puis appliquée à toutes les lignes
    try:
        raw_d1 = next((row.get('date_debut') for row in rows if row.get('date_debut') not in (None, '')), None)
        raw_d2 = next((row.get('date_fin') for row in rows if row.get('date_fin') not in (None, '')), None)
        if raw_d1 is None or str(raw_d1).strip() == '':
            issues.append(
                AnalysisIssue(
                    level='error',
                    row=2,
                    column='date_debut',
                    message='Date de début vide.',
                    how_to_fix='Renseignez date_debut dans la feuille Entête (ex. 13/07/2026).',
                )
            )
        else:
            date_debut = _parse_date(raw_d1, row=2, column='date_debut')
        if raw_d2 is None or str(raw_d2).strip() == '':
            issues.append(
                AnalysisIssue(
                    level='error',
                    row=2,
                    column='date_fin',
                    message='Date de fin vide.',
                    how_to_fix='Renseignez date_fin dans la feuille Entête (ex. 17/07/2026).',
                )
            )
        else:
            date_fin = _parse_date(raw_d2, row=2, column='date_fin')
    except ImportValidationError as exc:
        for err in exc.errors:
            issues.append(
                AnalysisIssue(
                    level='error',
                    row=err.get('row') or 2,
                    column=err.get('column'),
                    message=err.get('message') or str(exc),
                    how_to_fix=err.get('how_to_fix') or '',
                )
            )

    if date_debut and date_fin:
        if date_debut > date_fin:
            issues.append(
                AnalysisIssue(
                    level='error',
                    row=2,
                    column='date_debut',
                    message='La date de début est après la date de fin.',
                    how_to_fix='Inversez ou corrigez les deux dates dans la feuille Entête.',
                )
            )
        else:
            fixed_d1, fixed_d2, corrected = coerce_french_week_period(date_debut, date_fin, max_days=14)
            if corrected:
                date_debut, date_fin = fixed_d1, fixed_d2
                issues.append(
                    AnalysisIssue(
                        level='warning',
                        row=2,
                        column='date_debut',
                        message=(
                            f'Dates réinterprétées en jj/mm/aaaa : '
                            f'{date_debut.strftime("%d/%m/%Y")} → {date_fin.strftime("%d/%m/%Y")} '
                            f'(Excel avait lu une période impossible).'
                        ),
                        how_to_fix=(
                            'Préférez saisir les dates en texte jj/mm/aaaa dans Entête, '
                            'ou vérifier que Excel n’est pas en format mois/jour (US).'
                        ),
                    )
                )
            elif (date_fin - date_debut).days > 14:
                issues.append(
                    AnalysisIssue(
                        level='error',
                        row=2,
                        column='date_debut',
                        message=(
                            f'Période trop longue ({(date_fin - date_debut).days} jours) : '
                            f'{date_debut.strftime("%d/%m/%Y")} → {date_fin.strftime("%d/%m/%Y")}.'
                        ),
                        how_to_fix=(
                            'Un relevé hebdo fait au plus 14 jours. Vérifiez jj/mm/aaaa '
                            'dans la feuille Entête (ex. 03/08/2026 = 3 août, pas le 8 mars).'
                        ),
                    )
                )

    if date_debut and date_fin:
        for row in rows:
            row['date_debut'] = date_debut
            row['date_fin'] = date_fin

    for idx, row in enumerate(rows):
        excel_row = idx + 2
        cp_name = _to_name_or_none(row.get('id_cuve_principale'))
        cj_name = _to_name_or_none(row.get('id_cuve_journaliere'))
        raw_groupe = row.get('id_groupe')
        g_key = _groupe_key_from_raw(raw_groupe)
        if g_key and str(g_key).startswith('__NEW__'):
            g_key = None
        marque = _to_name_or_none(row.get('marque_groupe'))
        puissance_raw = row.get('puissance_groupe')
        puissance = (
            str(puissance_raw).strip()
            if puissance_raw is not None and str(puissance_raw).strip() != ''
            else None
        )

        # Pas d’id groupe : l’appli le déduit de marque + puissance
        if not g_key and (marque or puissance):
            existing = _lookup_groupe_by_marque_puissance(marque, puissance)
            if existing:
                g_key = existing.identifiant
                row['id_groupe'] = g_key
            else:
                g_key = _pending_groupe_key(marque, puissance)
                row['id_groupe'] = g_key
                row['marque_groupe'] = marque
                row['puissance_groupe'] = puissance

        for col in (
            'quantités_cuve_principale',
            'quantite_cuve_journaliere',
            'compteur_horaire',
            'depotage',
        ):
            try:
                val = _to_float(row.get(col), row=excel_row, column=col)
                if val < 0:
                    issues.append(
                        AnalysisIssue(
                            level='warning',
                            row=excel_row,
                            column=col,
                            message=f'Valeur négative ({val}) pour {col}.',
                            how_to_fix='Vérifiez le relevé : les volumes / compteurs sont en principe ≥ 0.',
                        )
                    )
            except ImportValidationError as exc:
                for err in exc.errors:
                    issues.append(
                        AnalysisIssue(
                            level='error',
                            row=err.get('row'),
                            column=err.get('column'),
                            message=err.get('message') or str(exc),
                            how_to_fix=err.get('how_to_fix') or '',
                        )
                    )

        if not any([cp_name, cj_name, g_key]):
            issues.append(
                AnalysisIssue(
                    level='warning',
                    row=excel_row,
                    column='id_cuve_principale',
                    message='Aucun site / cuve / groupe renseigné sur cette ligne.',
                    how_to_fix='Indiquez au moins le nom du site ou le n° de groupe.',
                )
            )

        key = (cp_name, cj_name, g_key)
        if any(key) and key in seen_keys:
            issues.append(
                AnalysisIssue(
                    level='warning',
                    row=excel_row,
                    column=None,
                    message=(
                        f'Ligne en double avec la ligne {seen_keys[key]} '
                        f'(même site / cuve journalière / groupe).'
                    ),
                    how_to_fix='Supprimez le doublon ou corrigez les identifiants.',
                )
            )
        elif any(key):
            seen_keys[key] = excel_row

        if cp_name:
            cp_names.setdefault(cp_name, []).append(excel_row)
        if cj_name:
            cj_names.setdefault(cj_name, []).append(excel_row)
            link = cj_links.setdefault(
                cj_name,
                {
                    'site': None,
                    'groupe_key': None,
                    'marque': None,
                    'puissance': None,
                    'capacite_cp': None,
                    'capacite_cj': None,
                },
            )
            if cp_name:
                link['site'] = cp_name
            if g_key:
                link['groupe_key'] = g_key
            if marque:
                link['marque'] = marque
            if puissance:
                link['puissance'] = puissance
            cap_cp = _optional_float(row.get('capacite_cp'))
            cap_cj = _optional_float(row.get('capacite_cj'))
            if cap_cp is not None:
                link['capacite_cp'] = cap_cp
            if cap_cj is not None:
                link['capacite_cj'] = cap_cj
        elif cp_name:
            # Site sans CJ encore : garder la capacité CP pour la création
            link = cj_links.setdefault(
                f'__site__{cp_name}',
                {
                    'site': cp_name,
                    'groupe_key': None,
                    'marque': None,
                    'puissance': None,
                    'capacite_cp': None,
                    'capacite_cj': None,
                },
            )
            cap_cp = _optional_float(row.get('capacite_cp'))
            if cap_cp is not None:
                link['capacite_cp'] = cap_cp
        if g_key:
            g_names.setdefault(g_key, []).append(excel_row)

    existing_cp = {}
    if cp_names:
        wanted = {n.strip().upper() for n in cp_names}
        for obj in CuvePrincipale.objects.exclude(identifiant='').only('identifiant'):
            if not obj.identifiant:
                continue
            key = obj.identifiant.strip().upper()
            if key in wanted:
                existing_cp[key] = obj.identifiant

    existing_cj = {}
    if cj_names:
        wanted = {n.strip().upper() for n in cj_names}
        for obj in CuveJournaliere.objects.exclude(identifiant__isnull=True).exclude(identifiant='').only('identifiant'):
            key = obj.identifiant.strip().upper()
            if key in wanted:
                existing_cj[key] = obj.identifiant

    existing_g: dict[str, str] = {}
    if g_names:
        wanted = {n.strip().upper() for n in g_names if not str(n).startswith('__NEW__')}
        for obj in GroupeElectrogene.objects.exclude(identifiant='').only('identifiant', 'pk'):
            if not obj.identifiant:
                continue
            key = obj.identifiant.strip().upper()
            if key in wanted:
                existing_g[key] = obj.identifiant
            pk_key = str(obj.pk)
            if pk_key.upper() in wanted and pk_key.upper() not in existing_g:
                existing_g[pk_key.upper()] = obj.identifiant

    def _split_names(names_map: dict[str, list[int]], existing_map: dict[str, str], kind: str, prefix: str):
        known, new = [], []
        for name, sample in sorted(names_map.items(), key=lambda x: x[0].upper()):
            if str(name).startswith('__NEW__'):
                ref = EntityRef(
                    kind=kind,
                    key=name,
                    label=f'{prefix} à créer depuis marque/puissance',
                    exists=False,
                    sample_rows=sample[:5],
                )
                new.append(ref)
                continue
            exists = name.strip().upper() in existing_map
            ref = EntityRef(
                kind=kind,
                key=name,
                label=f'{prefix} « {name} »',
                exists=exists,
                sample_rows=sample[:5],
            )
            (known if exists else new).append(ref)
        return known, new

    known_cp, new_cp = _split_names(cp_names, existing_cp, 'cuve_principale', 'Site')
    known_cj, new_cj = _split_names(cj_names, existing_cj, 'cuve_journaliere', 'Cuve journalière')
    known_g, new_g = _split_names(g_names, existing_g, 'groupe', 'Groupe')

    for ref in new_cp:
        issues.append(
            AnalysisIssue(
                level='warning',
                row=ref.sample_rows[0] if ref.sample_rows else None,
                column='id_cuve_principale',
                message=f'Nouveau site détecté : « {ref.key} ».',
                how_to_fix='Lancez create_rapport_entities pour le créer avant l’import des lignes.',
            )
        )
    for ref in new_cj:
        issues.append(
            AnalysisIssue(
                level='warning',
                row=ref.sample_rows[0] if ref.sample_rows else None,
                column='id_cuve_journaliere',
                message=f'Nouvelle cuve journalière détectée : « {ref.key} ».',
                how_to_fix='Lancez create_rapport_entities pour la créer avant l’import des lignes.',
            )
        )
    for ref in new_g:
        pending = _parse_pending_groupe_key(ref.key)
        if pending:
            marque, puissance = pending
            msg = f'Nouveau groupe à créer : marque « {marque} », puissance « {puissance} » → identifiant G{{n}}-{marque}-{puissance}.'
        else:
            msg = f'Nouveau groupe détecté : « {ref.key} ».'
        issues.append(
            AnalysisIssue(
                level='warning',
                row=ref.sample_rows[0] if ref.sample_rows else None,
                column='id_groupe',
                message=msg,
                how_to_fix='Lancez create_rapport_entities pour le créer avant l’import des lignes.',
            )
        )

    hard_errors = [i for i in issues if i.level == 'error']
    return AnalysisResult(
        ok=len(hard_errors) == 0,
        row_count=len(rows),
        date_debut=date_debut.isoformat() if date_debut else None,
        date_fin=date_fin.isoformat() if date_fin else None,
        issues=issues,
        known_cuves_principales=known_cp,
        known_cuves_journalieres=known_cj,
        known_groupes=known_g,
        new_cuves_principales=new_cp,
        new_cuves_journalieres=new_cj,
        new_groupes=new_g,
        cj_links=cj_links,
    )


def analyze_rapport_file(path: str | Path) -> AnalysisResult:
    return analyze_rapport_rows(load_rapport_rows(path))


def _optional_float(value: object) -> float | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '.').replace(' ', '')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _capacity_for_site(analysis: AnalysisResult, site_name: str | None, default: float) -> float:
    if not site_name:
        return default
    for link in analysis.cj_links.values():
        if link.get('site') == site_name and link.get('capacite_cp') is not None:
            return float(link['capacite_cp'])
    return default


def _capacity_for_cj(analysis: AnalysisResult, cj_name: str | None, default: float) -> float:
    if not cj_name:
        return default
    link = analysis.cj_links.get(cj_name) or {}
    if link.get('capacite_cj') is not None:
        return float(link['capacite_cj'])
    return default


@transaction.atomic
def create_entities_from_analysis(
    analysis: AnalysisResult,
    *,
    default_cp_capacity: float = 10000.0,
    default_cj_capacity: float = 1000.0,
    default_groupe_marque: str = 'À préciser',
    default_groupe_puissance: str = 'À préciser',
) -> CreateResult:
    result = CreateResult()

    for ref in analysis.new_cuves_principales:
        if _lookup_cp(ref.key):
            result.skipped_existing.append(f'site:{ref.key}')
            continue
        CuvePrincipale.objects.create(
            identifiant=ref.key.strip(),
            capacite=_capacity_for_site(analysis, ref.key, default_cp_capacity),
        )
        result.created_cuves_principales.append(ref.key)

    for ref in analysis.new_groupes:
        pending = _parse_pending_groupe_key(ref.key)
        if pending:
            marque, puissance = pending
            # Chercher infos plus riches dans cj_links
            for link in analysis.cj_links.values():
                if link.get('groupe_key') == ref.key:
                    marque = link.get('marque') or marque
                    puissance = link.get('puissance') or puissance
                    break
            existing = _lookup_groupe_by_marque_puissance(marque, puissance)
            if existing:
                result.skipped_existing.append(f'groupe:{existing.identifiant}')
                # Réécrire la clé pour les liaisons CJ
                for link in analysis.cj_links.values():
                    if link.get('groupe_key') == ref.key:
                        link['groupe_key'] = existing.identifiant
                continue
            identifiant = compose_groupe_identifiant(marque, puissance)
            GroupeElectrogene.objects.create(
                identifiant=identifiant,
                marque=(marque or 'À préciser').strip() or 'À préciser',
                puissance=str(puissance or '0').strip() or '0',
            )
            for link in analysis.cj_links.values():
                if link.get('groupe_key') == ref.key:
                    link['groupe_key'] = identifiant
            result.created_groupes.append(identifiant)
            continue

        identifiant = ref.key.strip()
        if GroupeElectrogene.objects.filter(identifiant__iexact=identifiant).exists():
            result.skipped_existing.append(f'groupe:{identifiant}')
            continue
        gid = _extract_groupe_id(identifiant)
        marque = default_groupe_marque
        puissance = default_groupe_puissance
        parts = identifiant.split('-')
        if len(parts) >= 3 and parts[0].upper().startswith('G'):
            # G1-SDMO-830
            marque = parts[1] if len(parts) > 1 else marque
            puissance = parts[2] if len(parts) > 2 else puissance
        elif len(parts) >= 2:
            if parts[-1].replace('.', '', 1).isdigit():
                puissance = parts[-1]
                marque = parts[-2]
            else:
                marque = parts[-1]
        defaults = {
            'identifiant': identifiant,
            'marque': marque,
            'puissance': str(puissance),
        }
        if gid is not None and not GroupeElectrogene.objects.filter(pk=gid).exists():
            GroupeElectrogene.objects.create(id=gid, **defaults)
        else:
            GroupeElectrogene.objects.create(**defaults)
        result.created_groupes.append(identifiant)

    for ref in analysis.new_cuves_journalieres:
        if _lookup_cj(ref.key):
            result.skipped_existing.append(f'cuve_journaliere:{ref.key}')
            continue
        link = analysis.cj_links.get(ref.key) or {}
        site_name = link.get('site')
        g_key = link.get('groupe_key') or link.get('groupe_id')

        if site_name and not _lookup_cp(site_name):
            CuvePrincipale.objects.create(
                identifiant=site_name.strip(),
                capacite=_capacity_for_site(analysis, site_name, default_cp_capacity),
            )
            if site_name not in result.created_cuves_principales:
                result.created_cuves_principales.append(site_name)

        cp = _lookup_cp(site_name) if site_name else None
        if not cp:
            raise ImportValidationError(
                f'Impossible de créer la cuve journalière « {ref.key} » : site manquant.',
                [
                    _friendly_error(
                        row=ref.sample_rows[0] if ref.sample_rows else None,
                        column='id_cuve_principale',
                        message=(
                            f'La cuve journalière « {ref.key} » nécessite un id_cuve_principale '
                            '(nom du site) sur la même ligne.'
                        ),
                        how_to_fix='Renseignez le nom du site comme sur la fiche de suivi.',
                    )
                ],
            )

        groupe = None
        if g_key is not None:
            if isinstance(g_key, int):
                groupe = _lookup_groupe(g_key)
            else:
                groupe = _lookup_groupe_by_identifiant(str(g_key))
        # OneToOne : détacher si le groupe est déjà lié ailleurs
        related_cj = getattr(groupe, 'cuve_journaliere', None) if groupe else None
        if related_cj is not None and getattr(related_cj, 'id', None) is not None:
            other = related_cj
            if other.identifiant != ref.key:
                other.groupe_electrogene = None
                other.save(update_fields=['groupe_electrogene'])

        CuveJournaliere.objects.create(
            identifiant=ref.key.strip(),
            capacite=_capacity_for_cj(analysis, ref.key, default_cj_capacity),
            cuve_principale=cp,
            groupe_electrogene=groupe,
        )
        result.created_cuves_journalieres.append(ref.key)

    return result


@transaction.atomic
def delete_rapport_and_orphans(rapport: Rapport) -> dict:
    """
    Supprime un rapport et les sites / cuves / groupes devenus orphelins
    (créés via un ajout de site et plus référencés par aucun autre rapport).
    """
    lignes = list(
        rapport.lignes.select_related(
            'cuve_principale',
            'cuve_journaliere',
            'groupe_electrogene',
        )
    )
    cp_ids = {l.cuve_principale_id for l in lignes if l.cuve_principale_id}
    cj_ids = {l.cuve_journaliere_id for l in lignes if l.cuve_journaliere_id}
    g_ids = {l.groupe_electrogene_id for l in lignes if l.groupe_electrogene_id}
    # Groupes liés aux CJ du rapport (même si pas sur la ligne)
    for cj_id in cj_ids:
        cj = CuveJournaliere.objects.filter(pk=cj_id).only('groupe_electrogene_id').first()
        if cj and cj.groupe_electrogene_id:
            g_ids.add(cj.groupe_electrogene_id)

    rapport_pk = rapport.id
    lignes_count = len(lignes)
    rapport.delete()

    deleted = {
        'rapport_id': rapport_pk,
        'lignes': lignes_count,
        'cuves_journalieres': [],
        'groupes': [],
        'cuves_principales': [],
    }

    for cj_id in cj_ids:
        still_used = LigneRapport.objects.filter(cuve_journaliere_id=cj_id).exists()
        if still_used:
            continue
        cj = CuveJournaliere.objects.filter(pk=cj_id).first()
        if not cj:
            continue
        name = cj.identifiant or str(cj_id)
        g = cj.groupe_electrogene
        if g:
            g_ids.add(g.id)
        cj.delete()
        deleted['cuves_journalieres'].append(name)

    for g_id in g_ids:
        still_used = (
            LigneRapport.objects.filter(groupe_electrogene_id=g_id).exists()
            or CuveJournaliere.objects.filter(groupe_electrogene_id=g_id).exists()
        )
        if still_used:
            continue
        g = GroupeElectrogene.objects.filter(pk=g_id).first()
        if not g:
            continue
        deleted['groupes'].append(g.identifiant)
        g.delete()

    for cp_id in cp_ids:
        still_used = (
            LigneRapport.objects.filter(cuve_principale_id=cp_id).exists()
            or CuveJournaliere.objects.filter(cuve_principale_id=cp_id).exists()
        )
        if still_used:
            continue
        cp = CuvePrincipale.objects.filter(pk=cp_id).first()
        if not cp:
            continue
        deleted['cuves_principales'].append(cp.identifiant)
        cp.delete()

    return deleted


@transaction.atomic
def import_rapport_lignes(
    rows: list[dict],
    user=None,
    *,
    create_missing: bool = False,
    require_entities: bool = True,
) -> ImportResult:
    rows = [normalize_row_keys(r) for r in rows]
    analysis = analyze_rapport_rows(rows, create_missing=create_missing)
    if not analysis.ok:
        raise ImportValidationError(
            (
                'Il y a 1 point à corriger dans votre fichier.'
                if len(analysis.errors) == 1
                else f'Il y a {len(analysis.errors)} points à corriger dans votre fichier.'
            ),
            [
                {
                    'row': i.row,
                    'column': i.column,
                    'column_label': i.column,
                    'message': i.message,
                    'how_to_fix': i.how_to_fix,
                }
                for i in analysis.errors[:30]
            ],
        )

    created_entities = None
    if create_missing and (
        analysis.new_cuves_principales
        or analysis.new_cuves_journalieres
        or analysis.new_groupes
    ):
        created_entities = create_entities_from_analysis(analysis)
        analysis = analyze_rapport_rows(rows, create_missing=False)

    if require_entities and (
        analysis.new_cuves_principales
        or analysis.new_cuves_journalieres
        or analysis.new_groupes
    ):
        missing_msgs = []
        for ref in analysis.new_cuves_principales:
            missing_msgs.append(
                _friendly_error(
                    row=ref.sample_rows[0] if ref.sample_rows else None,
                    column='id_cuve_principale',
                    message=f'Site inconnu : « {ref.key} ».',
                    how_to_fix='Exécutez : python manage.py create_rapport_entities <fichier>',
                )
            )
        for ref in analysis.new_cuves_journalieres:
            missing_msgs.append(
                _friendly_error(
                    row=ref.sample_rows[0] if ref.sample_rows else None,
                    column='id_cuve_journaliere',
                    message=f'Cuve journalière inconnue : « {ref.key} ».',
                    how_to_fix='Exécutez : python manage.py create_rapport_entities <fichier>',
                )
            )
        for ref in analysis.new_groupes:
            missing_msgs.append(
                _friendly_error(
                    row=ref.sample_rows[0] if ref.sample_rows else None,
                    column='id_groupe',
                    message=f'Groupe inconnu : « {ref.key} ».',
                    how_to_fix='Exécutez : python manage.py create_rapport_entities <fichier>',
                )
            )
        raise ImportValidationError(
            'Des sites / groupes du fichier n’existent pas encore en base.',
            missing_msgs[:30],
        )

    if not analysis.date_debut or not analysis.date_fin:
        raise ImportValidationError(
            'Les dates de période sont manquantes.',
            [
                _friendly_error(
                    row=2,
                    column='date_debut',
                    message='Impossible de lire la période du relevé.',
                    how_to_fix='Remplissez Date de début et Date de fin.',
                )
            ],
        )

    date_debut = date_cls.fromisoformat(analysis.date_debut)
    date_fin = date_cls.fromisoformat(analysis.date_fin)

    rapport = Rapport.objects.create(
        date_debut=date_debut,
        date_fin=date_fin,
        created_by=user if user is not None and getattr(user, 'pk', None) else None,
    )

    imported = 0
    for idx, row in enumerate(rows):
        excel_row = idx + 2
        cp_name = _to_name_or_none(row.get('id_cuve_principale'))
        cj_name = _to_name_or_none(row.get('id_cuve_journaliere'))
        raw_groupe = row.get('id_groupe')
        marque = _to_name_or_none(row.get('marque_groupe'))
        puissance_raw = row.get('puissance_groupe')
        puissance = (
            str(puissance_raw).strip()
            if puissance_raw is not None and str(puissance_raw).strip() != ''
            else None
        )
        cp = _lookup_cp(cp_name)
        cj = _lookup_cj(cj_name)
        g_key = _groupe_key_from_raw(raw_groupe)
        if g_key and str(g_key).startswith('__NEW__'):
            g_key = None
        groupe = _lookup_groupe_by_identifiant(g_key) if g_key else None
        if groupe is None and (marque or puissance):
            groupe = _lookup_groupe_by_marque_puissance(marque, puissance)

        LigneRapport.objects.create(
            rapport=rapport,
            cuve_principale=cp,
            cuve_journaliere=cj,
            groupe_electrogene=groupe,
            quantite_gasoil_cuve_principale=_to_float(
                row.get('quantités_cuve_principale'),
                row=excel_row,
                column='quantités_cuve_principale',
            ),
            quantite_gasoil_cuve_journaliere=_to_float(
                row.get('quantite_cuve_journaliere'),
                row=excel_row,
                column='quantite_cuve_journaliere',
            ),
            compteur_horaire=_to_float(
                row.get('compteur_horaire'),
                row=excel_row,
                column='compteur_horaire',
            ),
            depotage=_to_float(row.get('depotage'), row=excel_row, column='depotage'),
            etat_fonctionnement=str(row.get('état_fonctionnement') or 'F').strip() or 'F',
            observations=str(row.get('observations') or '').strip(),
        )
        imported += 1

    return ImportResult(
        rapport_id=rapport.id,
        imported_lines=imported,
        date_debut=analysis.date_debut,
        date_fin=analysis.date_fin,
        created_entities=created_entities,
    )


def import_rapport_file(
    path: str | Path,
    user=None,
    *,
    create_missing: bool = False,
    require_entities: bool = True,
) -> ImportResult:
    return import_rapport_lignes(
        load_rapport_rows(path),
        user=user,
        create_missing=create_missing,
        require_entities=require_entities,
    )
