import io
import tempfile
from pathlib import Path
from datetime import date
from django.core.management import call_command
from django.test import TestCase
from dashboard.models import CuvePrincipale, GroupeElectrogene, CuveJournaliere, Rapport, LigneRapport
from dashboard.norme import ImportValidationError, rows_from_csv, rows_from_xlsx
from dashboard.rapport_pipeline import (
    generate_rapport_template_xlsx,
    analyze_rapport_rows,
    import_rapport_lignes,
)
from dashboard.norme import build_rapport_csv_bytes, rapport_to_rows
from dashboard.sequence_utils import reset_sqlite_sequences


class WeeklyReportTemplateTestCase(TestCase):
    def setUp(self):
        self.cp = CuvePrincipale.objects.create(identifiant='SITE AKWA', capacite=10000.0)
        self.ge = GroupeElectrogene.objects.create(identifiant='1', marque='Perkins', puissance='100 kVA')
        self.cj = CuveJournaliere.objects.create(
            identifiant='CJ AKWA 1',
            capacite=1000.0,
            cuve_principale=self.cp,
            groupe_electrogene=self.ge,
        )

    def test_generate_template(self):
        content = generate_rapport_template_xlsx('2026-07-13', '2026-07-17')
        self.assertTrue(isinstance(content, bytes))
        self.assertTrue(len(content) > 1000)

        rows = rows_from_xlsx(content)
        self.assertTrue(len(rows) >= 1)
        first_row = rows[0]
        self.assertEqual(first_row.get('id_cuve_journaliere'), self.cj.identifiant)
        from dashboard.norme import _parse_date
        self.assertEqual(_parse_date(first_row.get('date_debut')), date(2026, 7, 13))
        self.assertEqual(_parse_date(first_row.get('date_fin')), date(2026, 7, 17))

    def test_analyze_generated_template(self):
        content = generate_rapport_template_xlsx('2026-07-13', '2026-07-17')
        rows = rows_from_xlsx(content)
        rows[0]['quantités_cuve_principale'] = 5000
        rows[0]['quantite_cuve_journaliere'] = 800
        rows[0]['compteur_horaire'] = 1200
        rows[0]['état_fonctionnement'] = 'F'

        analysis = analyze_rapport_rows(rows)
        self.assertTrue(analysis.ok)
        self.assertEqual(len(analysis.new_cuves_principales), 0)
        self.assertEqual(len(analysis.new_cuves_journalieres), 0)
        self.assertEqual(len(analysis.new_groupes), 0)

    def test_import_generated_template(self):
        content = generate_rapport_template_xlsx('2026-07-13', '2026-07-17')
        rows = rows_from_xlsx(content)
        rows[0]['quantités_cuve_principale'] = 4500
        rows[0]['quantite_cuve_journaliere'] = 750
        rows[0]['compteur_horaire'] = 1250
        rows[0]['état_fonctionnement'] = 'F'
        rows[0]['observations'] = 'Test OK'

        res = import_rapport_lignes(rows)
        self.assertEqual(res.imported_lines, 1)

        rapport = Rapport.objects.get(pk=res.rapport_id)
        self.assertEqual(rapport.date_debut, date(2026, 7, 13))
        self.assertEqual(rapport.date_fin, date(2026, 7, 17))

        ligne = rapport.lignes.first()
        self.assertEqual(ligne.cuve_journaliere, self.cj)
        self.assertEqual(ligne.cuve_principale, self.cp)
        self.assertEqual(ligne.groupe_electrogene, self.ge)
        self.assertEqual(ligne.compteur_horaire, 1250.0)

    def test_unknown_identifiant_blocks_import(self):
        content = generate_rapport_template_xlsx('2026-07-13', '2026-07-17')
        rows = rows_from_xlsx(content)
        rows[0]['id_cuve_journaliere'] = 'CUVE INCONNUE'

        analysis = analyze_rapport_rows(rows)
        self.assertTrue(analysis.ok)
        self.assertEqual(len(analysis.new_cuves_journalieres), 1)
        self.assertEqual(analysis.new_cuves_journalieres[0].key, 'CUVE INCONNUE')

    def test_import_creates_missing_entities_when_requested(self):
        content = generate_rapport_template_xlsx('2026-07-13', '2026-07-17')
        rows = rows_from_xlsx(content)
        rows[0]['id_cuve_journaliere'] = 'CJ NOUVELLE'
        rows[0]['id_cuve_principale'] = 'SITE AKWA'
        rows[0]['id_groupe'] = '1'
        rows[0]['quantités_cuve_principale'] = 4200
        rows[0]['quantite_cuve_journaliere'] = 700
        rows[0]['compteur_horaire'] = 1100
        rows[0]['état_fonctionnement'] = 'F'

        result = import_rapport_lignes(rows, create_missing=True)

        self.assertEqual(result.imported_lines, 1)
        created_cj = CuveJournaliere.objects.get(identifiant='CJ NOUVELLE')
        self.assertEqual(created_cj.cuve_principale.identifiant, 'SITE AKWA')
        self.assertEqual(created_cj.groupe_electrogene.identifiant, '1')

    def test_reset_sqlite_sequences_restores_report_ids_to_one(self):
        Rapport.objects.create(date_debut=date(2026, 7, 13), date_fin=date(2026, 7, 17))
        Rapport.objects.create(date_debut=date(2026, 7, 20), date_fin=date(2026, 7, 24))

        Rapport.objects.all().delete()
        reset_sqlite_sequences('dashboard')

        rapport = Rapport.objects.create(date_debut=date(2026, 7, 27), date_fin=date(2026, 7, 31))
        self.assertEqual(rapport.pk, 1)

    def test_import_csv_imports_reference_entities_before_report_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = Path(tmpdir)
            (data_dir / 'cuve_principale.csv').write_text(
                'id_cuve_principale,capcite\nSITE AKWA,10000\n',
                encoding='utf-8',
            )
            (data_dir / 'groupe_electrogene.csv').write_text(
                'id_groupe;marque_groupe;puissance_groupe\n1;Perkins;100 kVA\n',
                encoding='utf-8',
            )
            (data_dir / 'cuve_journaliere.csv').write_text(
                'id_cuve_journaliere,id_cuve_principale,capacite\nCJ AKWA 1,SITE AKWA,1000\n',
                encoding='utf-8',
            )
            (data_dir / 'cuve_journaliere_groupe.csv').write_text(
                'id_cuve_journaliere,id_groupe\nCJ AKWA 1,1\n',
                encoding='utf-8',
            )
            (data_dir / 'rapport_1_carburflow.csv').write_text(
                '# CARBURFLOW - RAPPORT IMPORTÉ\n'
                '# date_debut: 13/07/2026\n'
                '# date_fin: 17/07/2026\n'
                'id_cuve_journaliere,site,groupe_marque,quantite_cuve_principale,quantite_cuve_journaliere,depotage,compteur_horaire,etat_fonctionnement,observations\n'
                'CJ AKWA 1,SITE AKWA,1 (Perkins),4500.0,750.0,0.0,1250.0,F,Test OK\n',
                encoding='utf-8',
            )

            stdout = io.StringIO()
            stderr = io.StringIO()
            call_command('import_csv', dir=str(data_dir), stdout=stdout, stderr=stderr)

        self.assertTrue(CuvePrincipale.objects.filter(identifiant='SITE AKWA').exists())
        self.assertTrue(GroupeElectrogene.objects.filter(identifiant='1').exists())
        self.assertTrue(CuveJournaliere.objects.filter(identifiant='CJ AKWA 1').exists())
        self.assertTrue(Rapport.objects.exists())
        self.assertTrue(LigneRapport.objects.exists())

    def test_csv_import_accepts_metadata_comments(self):
        sample = (
            '# CARBURFLOW - RAPPORT IMPORTÉ\n'
            '# date_debut: 13/07/2026\n'
            '# date_fin: 17/07/2026\n'
            'id_cuve_journaliere,site,groupe_marque,quantite_cuve_principale,quantite_cuve_journaliere,depotage,compteur_horaire,etat_fonctionnement,observations\n'
            'CJ AKWA 1,SITE AKWA,1 (Perkins),4500.0,750.0,0.0,1250.0,F,Test OK\n'
        )

        rows = rows_from_csv(sample.encode('utf-8'))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['date_debut'], '13/07/2026')
        self.assertEqual(rows[0]['date_fin'], '17/07/2026')
        self.assertEqual(rows[0]['id_cuve_journaliere'], 'CJ AKWA 1')

    def test_export_uses_template_like_columns(self):
        rapport = Rapport.objects.create(date_debut=date(2026, 7, 13), date_fin=date(2026, 7, 17))
        LigneRapport.objects.create(
            rapport=rapport,
            cuve_principale=self.cp,
            cuve_journaliere=self.cj,
            groupe_electrogene=self.ge,
            quantite_gasoil_cuve_principale=4500.0,
            quantite_gasoil_cuve_journaliere=750.0,
            compteur_horaire=1250.0,
            depotage=0.0,
            etat_fonctionnement='F',
            observations='Test OK',
        )

        rows = rapport_to_rows(rapport)
        self.assertEqual(len(rows), 1)
        self.assertEqual(
            list(rows[0].keys()),
            [
                'id_cuve_journaliere',
                'site',
                'groupe_marque',
                'quantite_cuve_principale',
                'quantite_cuve_journaliere',
                'depotage',
                'compteur_horaire',
                'etat_fonctionnement',
                'observations',
            ],
        )
        self.assertEqual(rows[0]['id_cuve_journaliere'], self.cj.identifiant)
        self.assertEqual(rows[0]['site'], self.cp.identifiant)
        self.assertEqual(rows[0]['groupe_marque'], '1 (Perkins)')

        csv_bytes = build_rapport_csv_bytes(rapport)
        csv_text = csv_bytes.decode('utf-8-sig')
        self.assertIn('# CARBURFLOW — RAPPORT IMPORTÉ', csv_text)
        self.assertIn('# date_debut:', csv_text)
        self.assertIn('# date_fin:', csv_text)
        self.assertIn('id_cuve_journaliere', csv_text)

    def test_new_sites_section_does_not_shift_columns(self):
        """Ancien modèle Excel : zone nouveaux sites ≠ colonnes relevés."""
        from openpyxl import Workbook

        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = 'Entête'
        ws_meta.append(['date_debut', '20/07/2026', ''])
        ws_meta.append(['date_fin', '26/07/2026', ''])

        ws = wb.create_sheet('Relevés')
        ws.append([
            'Cuve journalière (Verrouillé)',
            'Site / Cuve Principale',
            'Groupe Électrogène',
            'Quantité CP (L)',
            'Quantité CJ (L)',
            'Dépotage (L)',
            'Compteur Horaire',
            'État (F/P/HS)',
            'Observations',
        ])
        ws.append([
            self.cj.identifiant,
            self.cp.identifiant,
            '1 (Perkins)',
            4500,
            750,
            0,
            1250,
            'F',
            'OK',
        ])
        ws.append(['--- NOUVEAUX SITES / GROUPES ---'])
        ws.append([
            'Cuve journalière (Laisser vide)',
            'Nom du Nouveau Site',
            'Code Site Existant (Opt.)',
            'Capacité CP (L)',
            'Capacité CJ (L)',
            'Marque Groupe',
            'Puissance Groupe',
            'Quantité CP (L)',
            'Quantité CJ (L)',
            'Dépotage (L)',
            'Compteur Horaire',
            'État (F/P/HS)',
            'Observations',
        ])
        # Col F = Marque SDMO, Col G = Puissance 830 — l’appli compose G{n}-SDMO-830
        ws.append([
            '',
            'NOUVEAU SITE TEST',
            '',
            12000,
            1500,
            'SDMO',
            '830',
            5000,
            800,
            100,
            42,
            'F',
            'Nouveau',
        ])

        out = io.BytesIO()
        wb.save(out)
        rows = rows_from_xlsx(out.getvalue())
        self.assertEqual(len(rows), 2)

        new_row = rows[1]
        self.assertEqual(new_row.get('id_cuve_principale'), 'NOUVEAU SITE TEST')
        self.assertEqual(new_row.get('id_cuve_journaliere'), 'NOUVEAU SITE TEST')
        self.assertEqual(new_row.get('marque_groupe'), 'SDMO')
        self.assertEqual(str(new_row.get('puissance_groupe')), '830')
        self.assertEqual(float(new_row.get('capacite_cp')), 12000.0)
        self.assertEqual(float(new_row.get('capacite_cj')), 1500.0)
        self.assertEqual(float(new_row.get('depotage')), 100.0)
        self.assertEqual(float(new_row.get('quantités_cuve_principale')), 5000.0)
        self.assertEqual(float(new_row.get('quantite_cuve_journaliere')), 800.0)

        analysis = analyze_rapport_rows(rows, create_missing=True)
        self.assertTrue(analysis.ok)
        result = import_rapport_lignes(rows, create_missing=True)
        self.assertEqual(result.imported_lines, 2)

        cp = CuvePrincipale.objects.get(identifiant='NOUVEAU SITE TEST')
        self.assertEqual(cp.capacite, 12000.0)
        cj = CuveJournaliere.objects.get(identifiant='NOUVEAU SITE TEST')
        self.assertEqual(cj.capacite, 1500.0)
        created = GroupeElectrogene.objects.filter(marque__iexact='SDMO', puissance='830').first()
        self.assertIsNotNone(created)
        self.assertRegex(created.identifiant, r'^G\d+-SDMO-830$')
        self.assertEqual(created.marque, 'SDMO')
        self.assertEqual(created.puissance, '830')

        ligne = LigneRapport.objects.filter(
            rapport_id=result.rapport_id,
            cuve_journaliere=cj,
        ).first()
        self.assertIsNotNone(ligne)
        self.assertEqual(ligne.quantite_gasoil_cuve_principale, 5000.0)
        self.assertEqual(ligne.quantite_gasoil_cuve_journaliere, 800.0)
        self.assertEqual(ligne.depotage, 100.0)
        self.assertEqual(ligne.compteur_horaire, 42.0)

    def test_delete_rapport_removes_orphaned_new_site(self):
        from dashboard.rapport_pipeline import delete_rapport_and_orphans

        rows = [
            {
                'date_debut': '20/07/2026',
                'date_fin': '26/07/2026',
                'id_cuve_principale': 'SITE ORPHELIN',
                'id_cuve_journaliere': 'SITE ORPHELIN',
                'marque_groupe': 'DEUTZ',
                'puissance_groupe': '450',
                'capacite_cp': 8000,
                'capacite_cj': 900,
                'quantités_cuve_principale': 3000,
                'quantite_cuve_journaliere': 400,
                'depotage': 0,
                'compteur_horaire': 10,
                'état_fonctionnement': 'F',
            }
        ]
        result = import_rapport_lignes(rows, create_missing=True)
        rapport = Rapport.objects.get(pk=result.rapport_id)
        self.assertTrue(CuvePrincipale.objects.filter(identifiant='SITE ORPHELIN').exists())

        deleted = delete_rapport_and_orphans(rapport)
        self.assertIn('SITE ORPHELIN', deleted['cuves_principales'])
        self.assertIn('SITE ORPHELIN', deleted['cuves_journalieres'])
        self.assertTrue(any('DEUTZ' in g for g in deleted['groupes']))
        self.assertFalse(CuvePrincipale.objects.filter(identifiant='SITE ORPHELIN').exists())
        self.assertFalse(CuveJournaliere.objects.filter(identifiant='SITE ORPHELIN').exists())
        self.assertTrue(CuvePrincipale.objects.filter(identifiant='SITE AKWA').exists())


class CalculsSeriesTestCase(TestCase):
    """Vérifie l'ordre chronologique et les trous (null) dans les séries."""

    def setUp(self):
        self.site = CuvePrincipale.objects.create(identifiant='SITE TEST', capacite=10000.0)
        self.groupe = GroupeElectrogene.objects.create(
            identifiant='G99-TEST-100', marque='TEST', puissance='100'
        )
        self.cj = CuveJournaliere.objects.create(
            identifiant='CJ TEST',
            capacite=2000.0,
            cuve_principale=self.site,
            groupe_electrogene=self.groupe,
        )
        self.r1 = Rapport.objects.create(date_debut=date(2026, 6, 22), date_fin=date(2026, 6, 26))
        self.r2 = Rapport.objects.create(date_debut=date(2026, 7, 6), date_fin=date(2026, 7, 10))
        self.r3 = Rapport.objects.create(date_debut=date(2026, 7, 13), date_fin=date(2026, 7, 17))

    def _line(self, rapport, volume_cp, volume_cj, compteur, depotage=0.0):
        return LigneRapport.objects.create(
            rapport=rapport,
            cuve_principale=self.site,
            cuve_journaliere=self.cj,
            groupe_electrogene=self.groupe,
            quantite_gasoil_cuve_principale=volume_cp,
            quantite_gasoil_cuve_journaliere=volume_cj,
            depotage=depotage,
            compteur_horaire=compteur,
        )

    def test_site_series_skips_missing_reports_without_fake_zero(self):
        from dashboard.utils import calculs as calc

        # Seulement r2 puis r3 (pas de ligne sur r1) — volume 3220 → 1440
        self._line(self.r2, 3000.0, 220.0, 100.0)
        self._line(self.r3, 1200.0, 240.0, 150.0)

        reports = list(Rapport.objects.order_by('date_debut', 'id'))
        lines_by_site_report = {}
        for line in LigneRapport.objects.all():
            lines_by_site_report.setdefault((self.site.id, line.rapport_id), []).append(line)

        volume, consumption = calc.calculer_site_series(reports, lines_by_site_report, self.site.id)

        # Volume site = CP réel (pas CP+CJ, pas somme multi-lignes)
        self.assertEqual(volume, [None, 3000.0, 1200.0])
        # Premier point présent = baseline 0 ; ensuite delta 3000 - 1200 = 1800
        self.assertEqual(consumption, [None, 0.0, 1800.0])

    def test_group_series_hours_and_consumption_follow_chrono(self):
        from dashboard.utils import calculs as calc

        self._line(self.r2, 3000.0, 220.0, 100.0)
        self._line(self.r3, 1200.0, 240.0, 150.0)

        reports = list(Rapport.objects.order_by('date_debut', 'id'))
        sites = [self.site]
        groupes = [self.groupe]

        lines_by_site_report = {}
        lines_by_group_report = {}
        groups_by_site_report = {}
        for line in LigneRapport.objects.all():
            lines_by_site_report.setdefault((self.site.id, line.rapport_id), []).append(line)
            lines_by_group_report.setdefault((self.groupe.id, line.rapport_id), []).append(line)
            groups_by_site_report.setdefault((self.site.id, line.rapport_id), set()).add(self.groupe.id)

        site_report_state = calc.build_site_report_state(reports, sites, lines_by_site_report)
        blocks = calc.calculer_groupes(
            reports=reports,
            groupes=groupes,
            sites=sites,
            lines_by_group_report=lines_by_group_report,
            site_report_state=site_report_state,
            groups_by_site_report=groups_by_site_report,
            groupes_by_id={self.groupe.id: self.groupe},
            group_primary_site_ids={self.groupe.id: self.site.id},
        )
        self.assertEqual(len(blocks), 1)
        block = blocks[0]
        self.assertEqual(block['volume'], [None, 3000.0, 1200.0])
        self.assertEqual(block['consumption'], [None, 0.0, 1800.0])
        # Premier compteur = baseline 0 h ; puis +50 h
        self.assertEqual(block['hours_run'], [None, 0.0, 50.0])
        self.assertAlmostEqual(block['latest_hourly_consumption'], 1800.0 / 50.0, places=2)
        # Autonomie = stock / conso horaire moyenne significative (même logique métriques)
        self.assertAlmostEqual(block['mean_hourly_consumption_deduite'], 36.0, places=2)
        self.assertAlmostEqual(block['volume_proportionnel'], 1440.0, places=1)
        self.assertAlmostEqual(block['autonomie_hours'], 1440.0 / 36.0, places=1)
        self.assertEqual(block['latest_main_volume'], 1200.0)
        self.assertEqual(block['latest_daily_volume'], 240.0)

    def test_autonomy_uses_power_share_on_shared_tank(self):
        from dashboard.utils import calculs as calc

        g2 = GroupeElectrogene.objects.create(
            identifiant='G98-TEST-200', marque='TEST', puissance='200'
        )
        CuveJournaliere.objects.create(
            identifiant='CJ TEST 2',
            capacite=2000.0,
            cuve_principale=self.site,
            groupe_electrogene=g2,
        )
        # G99 power 100, G98 power 200 → shares 1/3 and 2/3
        self._line(self.r3, 3000.0, 500.0, 50.0)
        LigneRapport.objects.create(
            rapport=self.r3,
            cuve_principale=self.site,
            cuve_journaliere=CuveJournaliere.objects.get(identifiant='CJ TEST 2'),
            groupe_electrogene=g2,
            quantite_gasoil_cuve_principale=3000.0,
            quantite_gasoil_cuve_journaliere=400.0,
            depotage=0,
            compteur_horaire=80.0,
        )
        # Baseline previous report for hours/consumption
        self._line(self.r2, 4000.0, 500.0, 10.0)
        LigneRapport.objects.create(
            rapport=self.r2,
            cuve_principale=self.site,
            cuve_journaliere=CuveJournaliere.objects.get(identifiant='CJ TEST 2'),
            groupe_electrogene=g2,
            quantite_gasoil_cuve_principale=4000.0,
            quantite_gasoil_cuve_journaliere=400.0,
            depotage=0,
            compteur_horaire=20.0,
        )

        reports = list(Rapport.objects.order_by('date_debut', 'id'))
        sites = [self.site]
        groupes = [self.groupe, g2]
        lines_by_site_report = {}
        lines_by_group_report = {}
        groups_by_site_report = {}
        for line in LigneRapport.objects.all():
            lines_by_site_report.setdefault((self.site.id, line.rapport_id), []).append(line)
            lines_by_group_report.setdefault((line.groupe_electrogene_id, line.rapport_id), []).append(line)
            groups_by_site_report.setdefault((self.site.id, line.rapport_id), set()).add(line.groupe_electrogene_id)

        blocks = calc.calculer_groupes(
            reports=reports,
            groupes=groupes,
            sites=sites,
            lines_by_group_report=lines_by_group_report,
            site_report_state=calc.build_site_report_state(reports, sites, lines_by_site_report),
            groups_by_site_report=groups_by_site_report,
            groupes_by_id={self.groupe.id: self.groupe, g2.id: g2},
            group_primary_site_ids={self.groupe.id: self.site.id, g2.id: self.site.id},
        )
        by_label = {b['label']: b for b in blocks}
        g100 = by_label['G99-TEST-100']
        # proportion 100/(100+200) = 1/3 → volume_prop = 3000/3 + 500 = 1500
        self.assertAlmostEqual(g100['power_share'], 100 / 300, places=4)
        self.assertEqual(g100['latest_main_volume'], 3000.0)
        self.assertEqual(g100['latest_daily_volume'], 500.0)
        self.assertAlmostEqual(g100['volume_proportionnel'], 1500.0, places=1)
        if g100['mean_hourly_consumption_deduite'] > 0:
            expected = g100['volume_proportionnel'] / g100['mean_hourly_consumption_deduite']
            self.assertAlmostEqual(g100['autonomie_hours'], expected, places=1)

    def test_site_volume_does_not_double_count_cp_across_groups(self):
        from dashboard.utils import calculs as calc

        g2 = GroupeElectrogene.objects.create(
            identifiant='G98-TEST-200', marque='TEST', puissance='200'
        )
        cj2 = CuveJournaliere.objects.create(
            identifiant='CJ TEST 2',
            capacite=2000.0,
            cuve_principale=self.site,
            groupe_electrogene=g2,
        )
        LigneRapport.objects.create(
            rapport=self.r2,
            cuve_principale=self.site,
            cuve_journaliere=self.cj,
            groupe_electrogene=self.groupe,
            quantite_gasoil_cuve_principale=5000.0,
            quantite_gasoil_cuve_journaliere=1000.0,
            depotage=0,
            compteur_horaire=10.0,
        )
        LigneRapport.objects.create(
            rapport=self.r2,
            cuve_principale=self.site,
            cuve_journaliere=cj2,
            groupe_electrogene=g2,
            quantite_gasoil_cuve_principale=5000.0,  # même CP répété
            quantite_gasoil_cuve_journaliere=800.0,
            depotage=0,
            compteur_horaire=20.0,
        )

        reports = [self.r2]
        lines_by_site_report = {}
        for line in LigneRapport.objects.filter(rapport=self.r2):
            lines_by_site_report.setdefault((self.site.id, line.rapport_id), []).append(line)

        volume, _ = calc.calculer_site_series(reports, lines_by_site_report, self.site.id)
        self.assertEqual(volume, [5000.0])


class RapportDateOrderTestCase(TestCase):
    def test_labels_are_chronological_start_date(self):
        from dashboard.utils import calculs as calc

        Rapport.objects.create(date_debut=date(2026, 7, 13), date_fin=date(2026, 7, 17))
        Rapport.objects.create(date_debut=date(2026, 6, 22), date_fin=date(2026, 6, 26))
        # Ancien bug Excel US : 03/08 stocké comme 08/03 — doit quand même se trier après correction
        Rapport.objects.create(date_debut=date(2026, 8, 3), date_fin=date(2026, 8, 9))

        reports = calc.ordered_rapports()
        labels = [calc.format_rapport_label(r) for r in reports]
        self.assertEqual(
            labels,
            [
                '22/06/2026',
                '13/07/2026',
                '03/08/2026',
            ],
        )

    def test_template_writes_text_french_dates(self):
        from openpyxl import load_workbook
        from dashboard.rapport_pipeline import generate_rapport_template_xlsx
        from dashboard.norme import _parse_date

        content = generate_rapport_template_xlsx('2026-08-03', '2026-08-09')
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb['Entête']
        # Ligne 4 = date_debut, ligne 5 = date_fin (après titre + vide + en-têtes)
        self.assertEqual(_parse_date(ws.cell(row=4, column=2).value), date(2026, 8, 3))
        self.assertEqual(_parse_date(ws.cell(row=5, column=2).value), date(2026, 8, 9))
        self.assertEqual(ws.cell(row=4, column=2).value, '03/08/2026')
        self.assertEqual(ws.cell(row=4, column=2).number_format, '@')

    def test_excel_us_swapped_debut_is_auto_corrected(self):
        """03/08 saisi (3 août) mais Excel US a stocké 08/03 (8 mars)."""
        from dashboard.rapport_pipeline import analyze_rapport_rows

        analysis = analyze_rapport_rows([
            {
                'date_debut': date(2026, 3, 8),  # lu US
                'date_fin': date(2026, 8, 11),   # correct 11 août
                'id_cuve_principale': 'SITE AKWA',
                'id_cuve_journaliere': 'CJ AKWA 1',
                'id_groupe': '1',
                'quantités_cuve_principale': 1000,
                'quantite_cuve_journaliere': 100,
                'depotage': 0,
                'compteur_horaire': 10,
                'état_fonctionnement': 'F',
            }
        ])
        self.assertEqual(analysis.date_debut, '2026-08-03')
        self.assertEqual(analysis.date_fin, '2026-08-11')
        self.assertFalse(any('trop longue' in (i.message or '') for i in analysis.issues))
        self.assertTrue(any('réinterprétées' in (i.message or '') for i in analysis.issues))

    def test_long_period_is_rejected(self):
        from dashboard.rapport_pipeline import analyze_rapport_rows

        # Période réellement trop longue, non récupérable par échange jj/mm
        analysis = analyze_rapport_rows([
            {
                'date_debut': '01/01/2026',
                'date_fin': '01/08/2026',
                'id_cuve_principale': 'SITE AKWA',
                'id_cuve_journaliere': 'CJ AKWA 1',
                'id_groupe': '1',
                'quantités_cuve_principale': 1000,
                'quantite_cuve_journaliere': 100,
                'depotage': 0,
                'compteur_horaire': 10,
                'état_fonctionnement': 'F',
            }
        ])
        self.assertFalse(analysis.ok)
        self.assertTrue(any('trop longue' in (i.message or '') for i in analysis.issues))
